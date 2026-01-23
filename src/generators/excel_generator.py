"""
excel_generator.py - 네이버 스마트스토어 대량등록용 엑셀 생성기

Phase 6-1: Gemini CTO 권장 - "자동 등록 API 대신 엑셀 내보내기"

기능:
- 분석 결과를 네이버 대량등록 양식으로 변환
- .xlsx 파일 생성
- 사용자는 파일만 받아 네이버에 업로드

네이버 대량등록 가이드:
https://help.sell.smartstore.naver.com/seller/faq/CONTENTS_FNQ
"""

import os
from datetime import datetime
from dataclasses import dataclass
from typing import List, Optional, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Dummy classes for type hints
    openpyxl = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None


@dataclass
class NaverProductData:
    """네이버 스마트스토어 상품 등록용 데이터"""
    # 필수 항목
    product_name: str                    # 상품명 (최대 100자)
    sale_price: int                      # 판매가
    stock_quantity: int = 999            # 재고수량 (기본 999)

    # 선택 항목
    category_id: str = ""                # 카테고리 ID
    brand: str = ""                      # 브랜드
    manufacturer: str = ""               # 제조사
    origin: str = "중국"                 # 원산지

    # 이미지
    main_image_url: str = ""             # 대표 이미지 URL
    detail_images: List[str] = None      # 추가 이미지 URLs

    # 배송 정보
    shipping_fee: int = 3000             # 배송비
    shipping_method: str = "택배"        # 배송방법

    # 상세 설명
    detail_content: str = ""             # 상품 상세 (HTML)

    # 옵션
    options: List[Dict[str, Any]] = None # 옵션 정보

    # 마진 분석 결과 (내부용)
    cost_price: int = 0                  # 원가
    margin_rate: float = 0.0             # 마진율
    breakeven_price: int = 0             # 손익분기가
    risk_level: str = ""                 # 위험도

    # 소싱 정보
    source_url: str = ""                 # 1688 URL
    source_price_cny: float = 0.0        # 1688 가격 (위안)
    moq: int = 1                         # MOQ


class NaverExcelGenerator:
    """네이버 스마트스토어 대량등록 엑셀 생성기

    Usage:
        generator = NaverExcelGenerator()
        products = [NaverProductData(...), ...]
        filepath = generator.generate(products, "output/products.xlsx")
    """

    # 네이버 대량등록 양식 컬럼 (간소화 버전)
    COLUMNS = [
        ("A", "상품명", 40),
        ("B", "판매가", 12),
        ("C", "재고수량", 10),
        ("D", "카테고리ID", 15),
        ("E", "브랜드", 15),
        ("F", "제조사", 15),
        ("G", "원산지", 10),
        ("H", "대표이미지URL", 50),
        ("I", "배송비", 10),
        ("J", "상품상세", 60),
        # 내부 분석용 (네이버 업로드 시 삭제)
        ("K", "[분석]원가", 12),
        ("L", "[분석]마진율", 10),
        ("M", "[분석]손익분기가", 12),
        ("N", "[분석]위험도", 10),
        ("O", "[분석]1688URL", 50),
        ("P", "[분석]1688가격(CNY)", 15),
        ("Q", "[분석]MOQ", 8),
    ]

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl 패키지가 필요합니다.\n"
                "설치: pip install openpyxl"
            )

        # 스타일 정의 (openpyxl import 후 초기화)
        self.HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        self.ANALYSIS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.RISK_COLORS = {
            "safe": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "danger": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

    def generate(
        self,
        products: List[NaverProductData],
        output_path: str = None,
        include_analysis: bool = True
    ) -> str:
        """엑셀 파일 생성

        Args:
            products: 상품 데이터 리스트
            output_path: 출력 파일 경로 (없으면 자동 생성)
            include_analysis: 분석 컬럼 포함 여부

        Returns:
            생성된 파일 경로
        """
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"output/naver_products_{timestamp}.xlsx"

        # 디렉토리 생성
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "상품목록"

        # 헤더 작성
        self._write_header(ws, include_analysis)

        # 데이터 작성
        for idx, product in enumerate(products, start=2):
            self._write_product_row(ws, idx, product, include_analysis)

        # 열 너비 조정
        self._adjust_column_widths(ws, include_analysis)

        # 필터 추가
        ws.auto_filter.ref = ws.dimensions

        # 사용 가이드 시트 추가
        self._add_guide_sheet(wb)

        # 저장
        wb.save(output_path)
        print(f"✅ 엑셀 파일 생성 완료: {output_path}")
        print(f"   - 상품 수: {len(products)}개")

        return output_path

    def _write_header(self, ws, include_analysis: bool):
        """헤더 행 작성"""
        columns = self.COLUMNS if include_analysis else self.COLUMNS[:10]

        for col_letter, col_name, _ in columns:
            cell = ws[f"{col_letter}1"]
            cell.value = col_name
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 분석 컬럼은 다른 색상
            if col_name.startswith("[분석]"):
                cell.fill = self.ANALYSIS_FILL
            else:
                cell.fill = self.HEADER_FILL

        # 헤더 행 고정
        ws.freeze_panes = "A2"

    def _write_product_row(self, ws, row: int, product: NaverProductData, include_analysis: bool):
        """상품 데이터 행 작성"""
        # 기본 정보
        ws[f"A{row}"] = product.product_name[:100]  # 최대 100자
        ws[f"B{row}"] = product.sale_price
        ws[f"C{row}"] = product.stock_quantity
        ws[f"D{row}"] = product.category_id
        ws[f"E{row}"] = product.brand
        ws[f"F{row}"] = product.manufacturer
        ws[f"G{row}"] = product.origin
        ws[f"H{row}"] = product.main_image_url
        ws[f"I{row}"] = product.shipping_fee
        ws[f"J{row}"] = product.detail_content[:500] if product.detail_content else ""

        # 분석 정보 (선택)
        if include_analysis:
            ws[f"K{row}"] = product.cost_price
            ws[f"L{row}"] = f"{product.margin_rate:.1f}%"
            ws[f"M{row}"] = product.breakeven_price

            # 위험도 셀 색상
            risk_cell = ws[f"N{row}"]
            risk_cell.value = product.risk_level
            if product.risk_level.lower() in self.RISK_COLORS:
                risk_cell.fill = self.RISK_COLORS[product.risk_level.lower()]

            ws[f"O{row}"] = product.source_url
            ws[f"P{row}"] = product.source_price_cny
            ws[f"Q{row}"] = product.moq

    def _adjust_column_widths(self, ws, include_analysis: bool):
        """열 너비 조정"""
        columns = self.COLUMNS if include_analysis else self.COLUMNS[:10]

        for col_letter, _, width in columns:
            ws.column_dimensions[col_letter].width = width

    def _add_guide_sheet(self, wb):
        """사용 가이드 시트 추가"""
        ws = wb.create_sheet("사용가이드")

        guide_text = [
            ("네이버 스마트스토어 대량등록 가이드", True),
            ("", False),
            ("1. 상품목록 시트의 A~J 컬럼만 네이버에 업로드합니다.", False),
            ("2. K~Q 컬럼([분석]으로 시작)은 내부 분석용이므로 업로드 전 삭제하세요.", False),
            ("3. 이미지 URL은 네이버에서 접근 가능한 URL이어야 합니다.", False),
            ("4. 카테고리ID는 네이버 판매자센터에서 확인하세요.", False),
            ("", False),
            ("위험도 설명:", True),
            ("- safe (녹색): 마진율 30% 이상 - 안전", False),
            ("- warning (노란색): 마진율 15~30% - 주의", False),
            ("- danger (빨간색): 마진율 15% 미만 - 위험 (판매 비추천)", False),
            ("", False),
            ("문의: https://github.com/hyunwoooim-star/smart-store-agent", False),
        ]

        for idx, (text, is_header) in enumerate(guide_text, start=1):
            cell = ws[f"A{idx}"]
            cell.value = text
            if is_header:
                cell.font = Font(bold=True, size=12)

        ws.column_dimensions["A"].width = 80


def create_naver_product_from_analysis(
    product_name: str,
    sale_price: int,
    cost_result: Any,  # CostResult
    scraped_product: Any = None,  # ScrapedProduct
    category_id: str = "",
) -> NaverProductData:
    """분석 결과를 NaverProductData로 변환

    Args:
        product_name: 상품명
        sale_price: 판매가
        cost_result: 마진 계산 결과 (CostResult)
        scraped_product: 스크래핑 결과 (ScrapedProduct, 선택)
        category_id: 네이버 카테고리 ID

    Returns:
        NaverProductData: 네이버 등록용 데이터
    """
    data = NaverProductData(
        product_name=product_name,
        sale_price=sale_price,
        category_id=category_id,
        cost_price=cost_result.total_cost,
        margin_rate=cost_result.margin_percent,
        breakeven_price=cost_result.breakeven_price,
        risk_level=cost_result.risk_level.value if hasattr(cost_result.risk_level, 'value') else str(cost_result.risk_level),
    )

    # 스크래핑 데이터가 있으면 추가
    if scraped_product:
        data.source_url = getattr(scraped_product, 'url', '')
        data.source_price_cny = getattr(scraped_product, 'price_cny', 0.0)
        data.moq = getattr(scraped_product, 'moq', 1)
        data.main_image_url = getattr(scraped_product, 'image_url', '') or ''

    return data


# CLI 테스트용
if __name__ == "__main__":
    # Mock 데이터로 테스트
    test_products = [
        NaverProductData(
            product_name="초경량 캠핑의자 접이식 릴렉스체어",
            sale_price=45000,
            category_id="50000803",
            origin="중국",
            main_image_url="https://example.com/chair.jpg",
            shipping_fee=3000,
            cost_price=28500,
            margin_rate=25.5,
            breakeven_price=35000,
            risk_level="warning",
            source_url="https://detail.1688.com/offer/123456.html",
            source_price_cny=45.0,
            moq=50,
        ),
        NaverProductData(
            product_name="휴대용 LED 캠핑랜턴 충전식",
            sale_price=25000,
            category_id="50000804",
            origin="중국",
            main_image_url="https://example.com/lantern.jpg",
            shipping_fee=3000,
            cost_price=12000,
            margin_rate=42.0,
            breakeven_price=15000,
            risk_level="safe",
            source_url="https://detail.1688.com/offer/789012.html",
            source_price_cny=25.0,
            moq=100,
        ),
    ]

    generator = NaverExcelGenerator()
    filepath = generator.generate(test_products, "output/test_naver_products.xlsx")
    print(f"\n테스트 파일 생성: {filepath}")
