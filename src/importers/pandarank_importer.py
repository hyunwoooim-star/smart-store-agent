"""
pandarank_importer.py - 판다랭크/아이템스카우트 엑셀 가져오기

Gemini CTO 제안:
"판다랭크에서 뽑은 꿀키워드 50개, 오늘 밤에 봇이 1688에서 싹 뒤져놓음."

워크플로우:
1. 판다랭크에서 키워드 엑셀 다운로드
2. 이 모듈로 엑셀 업로드
3. Night Crawler 큐에 자동 추가
4. 다음 날 아침 마진 분석 결과 확인

지원 형식:
- 판다랭크 키워드 분석 엑셀
- 아이템스카우트 엑셀
- 일반 키워드 목록 (1열)
"""

import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
import re

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ImportedKeyword:
    """가져온 키워드 데이터"""
    keyword: str
    search_volume: int = 0          # 검색량
    competition_rate: float = 0.0   # 경쟁 강도 (0~1)
    shopping_rate: float = 0.0      # 쇼핑성 지수
    category: str = ""              # 카테고리
    avg_price: int = 0              # 평균 가격
    priority: int = 5               # 우선순위 (1~10)


class PandarankImporter:
    """판다랭크/아이템스카우트 엑셀 가져오기"""

    # 판다랭크 컬럼 매핑 (다양한 형식 지원)
    COLUMN_MAPPINGS = {
        "keyword": ["키워드", "검색어", "keyword", "검색 키워드"],
        "search_volume": ["검색량", "월간검색량", "search_volume", "검색수"],
        "competition": ["경쟁강도", "경쟁률", "competition", "경쟁지수", "경쟁 강도"],
        "shopping_rate": ["쇼핑성", "쇼핑지수", "shopping_rate", "쇼핑성 지수"],
        "category": ["카테고리", "category", "분류"],
        "avg_price": ["평균가", "평균가격", "avg_price", "시장가"],
    }

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 패키지가 필요합니다: pip install openpyxl")

    def import_excel(self, file_path: str) -> Tuple[List[ImportedKeyword], Dict]:
        """엑셀 파일에서 키워드 가져오기

        Args:
            file_path: 엑셀 파일 경로

        Returns:
            Tuple[List[ImportedKeyword], Dict]: (키워드 목록, 통계)
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

        wb = load_workbook(path, read_only=True)
        ws = wb.active

        # 헤더 분석
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        column_map = self._detect_columns(headers)

        # 키워드 컬럼 필수 체크 제거 - _detect_columns에서 첫 컬럼 자동 사용
        if "keyword" not in column_map:
            # 이 경우는 헤더가 아예 없는 경우
            raise ValueError("엑셀 파일이 비어있거나 형식이 올바르지 않습니다.")

        # 데이터 읽기
        keywords = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # 빈 행 스킵
                continue

            keyword_val = str(row[column_map["keyword"]] or "").strip()
            if not keyword_val or len(keyword_val) < 2:
                continue

            # 키워드 정제
            keyword_val = self._clean_keyword(keyword_val)
            if not keyword_val:
                continue

            kw = ImportedKeyword(keyword=keyword_val)

            # 선택적 컬럼 읽기
            if "search_volume" in column_map:
                val = row[column_map["search_volume"]]
                kw.search_volume = self._parse_int(val)

            if "competition" in column_map:
                val = row[column_map["competition"]]
                kw.competition_rate = self._parse_float(val)

            if "shopping_rate" in column_map:
                val = row[column_map["shopping_rate"]]
                kw.shopping_rate = self._parse_float(val)

            if "category" in column_map:
                val = row[column_map["category"]]
                kw.category = str(val or "").strip()

            if "avg_price" in column_map:
                val = row[column_map["avg_price"]]
                kw.avg_price = self._parse_int(val)

            # 우선순위 계산 (검색량 높고 경쟁 낮으면 높은 점수)
            kw.priority = self._calculate_priority(kw)

            keywords.append(kw)

        wb.close()

        # 통계 계산
        stats = self._calculate_stats(keywords)

        return keywords, stats

    def import_from_bytes(self, file_bytes: bytes, filename: str = "upload.xlsx") -> Tuple[List[ImportedKeyword], Dict]:
        """바이트 데이터에서 키워드 가져오기 (Streamlit 업로드용)

        Args:
            file_bytes: 파일 바이트 데이터
            filename: 파일명

        Returns:
            Tuple[List[ImportedKeyword], Dict]: (키워드 목록, 통계)
        """
        import tempfile
        import os

        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        try:
            return self.import_excel(tmp_path)
        finally:
            os.unlink(tmp_path)

    def _detect_columns(self, headers: List[str]) -> Dict[str, int]:
        """헤더에서 컬럼 위치 감지

        Args:
            headers: 헤더 목록

        Returns:
            Dict: 컬럼명 → 인덱스 매핑
        """
        column_map = {}
        headers_lower = [h.lower() for h in headers]

        for field, aliases in self.COLUMN_MAPPINGS.items():
            for i, header in enumerate(headers_lower):
                for alias in aliases:
                    if alias.lower() in header:
                        column_map[field] = i
                        break
                if field in column_map:
                    break

        # 키워드 컬럼을 못 찾으면 첫 번째 컬럼을 키워드로 가정
        if "keyword" not in column_map and headers:
            column_map["keyword"] = 0
            print(f"[Importer] 키워드 컬럼 자동 감지: 첫 번째 컬럼 '{headers[0]}' 사용")

        return column_map

    def _clean_keyword(self, keyword: str) -> str:
        """키워드 정제

        Args:
            keyword: 원본 키워드

        Returns:
            str: 정제된 키워드
        """
        # 특수문자 제거 (공백, 한글, 영문, 숫자만 유지)
        keyword = re.sub(r'[^\w\s가-힣]', '', keyword)
        # 공백 정규화
        keyword = ' '.join(keyword.split())
        return keyword.strip()

    def _parse_int(self, value) -> int:
        """정수 파싱"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        try:
            # 문자열에서 숫자만 추출
            cleaned = re.sub(r'[^\d]', '', str(value))
            return int(cleaned) if cleaned else 0
        except:
            return 0

    def _parse_float(self, value) -> float:
        """실수 파싱"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            # 문자열에서 숫자와 점만 추출
            cleaned = re.sub(r'[^\d.]', '', str(value))
            return float(cleaned) if cleaned else 0.0
        except:
            return 0.0

    def _calculate_priority(self, kw: ImportedKeyword) -> int:
        """우선순위 계산 (1~10)

        높은 검색량 + 낮은 경쟁 + 높은 쇼핑성 = 높은 우선순위

        Args:
            kw: 키워드 데이터

        Returns:
            int: 우선순위 (1~10)
        """
        score = 5.0  # 기본값

        # 검색량 점수 (0~3점)
        if kw.search_volume >= 10000:
            score += 3
        elif kw.search_volume >= 5000:
            score += 2
        elif kw.search_volume >= 1000:
            score += 1

        # 경쟁 점수 (0~3점, 낮을수록 좋음)
        if kw.competition_rate <= 0.3:
            score += 3
        elif kw.competition_rate <= 0.5:
            score += 2
        elif kw.competition_rate <= 0.7:
            score += 1

        # 쇼핑성 점수 (0~2점)
        if kw.shopping_rate >= 0.7:
            score += 2
        elif kw.shopping_rate >= 0.5:
            score += 1

        return min(10, max(1, int(score)))

    def _calculate_stats(self, keywords: List[ImportedKeyword]) -> Dict:
        """통계 계산

        Args:
            keywords: 키워드 목록

        Returns:
            Dict: 통계 정보
        """
        if not keywords:
            return {
                "total": 0,
                "avg_search_volume": 0,
                "avg_competition": 0,
                "high_priority_count": 0,
            }

        return {
            "total": len(keywords),
            "avg_search_volume": sum(k.search_volume for k in keywords) // len(keywords),
            "avg_competition": sum(k.competition_rate for k in keywords) / len(keywords),
            "high_priority_count": len([k for k in keywords if k.priority >= 7]),
            "categories": list(set(k.category for k in keywords if k.category)),
        }


def add_keywords_to_crawler(keywords: List[ImportedKeyword], keyword_manager) -> int:
    """가져온 키워드를 Night Crawler 큐에 추가

    Args:
        keywords: 가져온 키워드 목록
        keyword_manager: KeywordManager 인스턴스

    Returns:
        int: 추가된 키워드 수
    """
    added = 0
    for kw in keywords:
        try:
            keyword_manager.add_keyword(
                keyword=kw.keyword,
                category=kw.category or "일반",
                priority=kw.priority
            )
            added += 1
        except Exception as e:
            print(f"[Importer] 키워드 추가 실패: {kw.keyword} - {e}")

    return added


# CLI 테스트
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("사용법: python pandarank_importer.py <엑셀파일경로>")
        sys.exit(1)

    importer = PandarankImporter()
    keywords, stats = importer.import_excel(sys.argv[1])

    print(f"\n=== 가져오기 완료 ===")
    print(f"총 키워드: {stats['total']}개")
    print(f"평균 검색량: {stats['avg_search_volume']:,}")
    print(f"평균 경쟁강도: {stats['avg_competition']:.2f}")
    print(f"고우선순위(7+): {stats['high_priority_count']}개")

    print(f"\n상위 10개 키워드:")
    for kw in sorted(keywords, key=lambda x: x.priority, reverse=True)[:10]:
        print(f"  [{kw.priority}] {kw.keyword} (검색량: {kw.search_volume:,})")
