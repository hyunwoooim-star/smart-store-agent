"""
naver_excel_exporter.py - 네이버 대량등록용 엑셀 추출기

워크플로우:
    Night Crawler → 모닝 브리핑 승인 → 엑셀 다운로드 → 네이버 대량등록

사용법:
    exporter = NaverExcelExporter()
    file_path = exporter.export_approved()
"""

import os
from datetime import datetime
from typing import List, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.crawler_models import SourcingCandidate, CandidateStatus
from src.crawler.repository import CandidateRepository
from src.publisher.content_generator import ContentGenerator


class NaverExcelExporter:
    """네이버 대량등록용 엑셀 추출기"""

    # 네이버 대량등록 필수 컬럼
    COLUMNS = [
        "상품명",           # A
        "판매가",           # B
        "원가(참고용)",      # C
        "마진율",           # D
        "카테고리",         # E
        "상품상태",         # F
        "배송비 유형",       # G
        "배송비",           # H
        "재고수량",         # I
        "대표이미지URL",     # J
        "추가이미지URL",     # K
        "상세페이지HTML",    # L
        "1688 URL",        # M
        "메모",            # N
    ]

    def __init__(
        self,
        repository: CandidateRepository = None,
        content_generator: ContentGenerator = None,
        output_dir: str = "data/exports"
    ):
        self.repository = repository or CandidateRepository()
        self.content_generator = content_generator or ContentGenerator()
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_approved(self, filename: str = None) -> str:
        """승인된 상품을 엑셀로 추출

        Args:
            filename: 출력 파일명 (없으면 자동 생성)

        Returns:
            str: 생성된 엑셀 파일 경로
        """
        # 승인된 후보만 가져오기
        candidates = self.repository.get_candidates_by_status(CandidateStatus.APPROVED)

        if not candidates:
            print("[ExcelExporter] 승인된 상품이 없습니다.")
            return None

        print(f"[ExcelExporter] {len(candidates)}개 상품 추출 중...")

        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "네이버 대량등록"

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 추가
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 데이터 추가
        for row, candidate in enumerate(candidates, 2):
            self._write_candidate_row(ws, row, candidate)

        # 열 너비 자동 조정
        self._adjust_column_widths(ws)

        # 파일 저장
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"naver_bulk_{timestamp}.xlsx"

        filepath = self.output_dir / filename
        wb.save(filepath)

        print(f"[ExcelExporter] 저장 완료: {filepath}")
        return str(filepath)

    def _write_candidate_row(self, ws, row: int, candidate: SourcingCandidate):
        """후보 상품 데이터를 행에 작성

        Args:
            ws: 워크시트
            row: 행 번호
            candidate: 소싱 후보
        """
        # 상세페이지 HTML 생성
        detail_html = self._generate_detail_html(candidate)

        # 추가 이미지 (첫 번째 제외)
        extra_images = ""
        if len(candidate.source_images) > 1:
            extra_images = "\n".join(candidate.source_images[1:5])  # 최대 4개

        # 메모 생성
        memo = self._generate_memo(candidate)

        # 데이터 작성
        data = [
            candidate.title_kr,                                    # 상품명
            candidate.recommended_price,                           # 판매가
            int(candidate.estimated_cost_krw),                     # 원가
            f"{candidate.estimated_margin_rate:.0%}",              # 마진율
            candidate.keyword or "생활용품",                        # 카테고리
            "새상품",                                               # 상품상태
            "무료",                                                 # 배송비 유형
            0,                                                     # 배송비
            999,                                                   # 재고수량
            candidate.source_images[0] if candidate.source_images else "",  # 대표이미지
            extra_images,                                          # 추가이미지
            detail_html,                                           # 상세페이지
            candidate.source_url,                                  # 1688 URL
            memo,                                                  # 메모
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)

            # 마진율 셀 스타일 (40% 이상이면 녹색)
            if col == 4 and candidate.estimated_margin_rate >= 0.40:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def _generate_detail_html(self, candidate: SourcingCandidate) -> str:
        """상세페이지 HTML 생성 (PAS 프레임워크)

        Args:
            candidate: 소싱 후보

        Returns:
            str: HTML 코드
        """
        # 간단한 PAS 템플릿
        html = f"""
<div style="max-width: 860px; margin: 0 auto; padding: 20px; font-family: 'Noto Sans KR', sans-serif;">

    <!-- 헤드카피 -->
    <h1 style="text-align: center; font-size: 28px; margin-bottom: 30px;">
        {candidate.title_kr}
    </h1>

    <!-- 대표 이미지 -->
    <div style="text-align: center; margin-bottom: 40px;">
        <img src="{candidate.source_images[0] if candidate.source_images else ''}"
             style="max-width: 100%; height: auto;"
             alt="{candidate.title_kr}">
    </div>

    <!-- PAS: Problem -->
    <div style="background: #FFF3E0; padding: 20px; margin: 20px 0; border-radius: 8px;">
        <h2 style="color: #E65100;">❗ 이런 고민 있으시죠?</h2>
        <ul style="line-height: 2;">
            <li>시중 제품은 품질 대비 가격이 너무 비싸다</li>
            <li>좋은 제품을 찾기 위해 시간을 너무 많이 쓴다</li>
            <li>막상 사면 기대에 못 미치는 경우가 많다</li>
        </ul>
    </div>

    <!-- PAS: Agitation -->
    <div style="background: #FCE4EC; padding: 20px; margin: 20px 0; border-radius: 8px;">
        <h2 style="color: #C2185B;">⚠️ 이대로 괜찮을까요?</h2>
        <p style="line-height: 1.8;">
            저렴한 제품을 사면 금방 망가지고, 비싼 제품을 사면 가성비가 떨어지고...<br>
            결국 돈도 시간도 낭비하게 됩니다.
        </p>
    </div>

    <!-- PAS: Solution -->
    <div style="background: #E8F5E9; padding: 20px; margin: 20px 0; border-radius: 8px;">
        <h2 style="color: #388E3C;">✅ 해결책을 찾았습니다!</h2>
        <p style="line-height: 1.8;">
            <strong>{candidate.title_kr}</strong>은(는) 중국 공장 직접 소싱으로<br>
            <strong>중간 마진 없이</strong> 합리적인 가격에 제공됩니다.
        </p>
        <p style="font-size: 24px; text-align: center; color: #1B5E20; margin: 20px 0;">
            🏷️ 판매가: <strong>{candidate.recommended_price:,}원</strong>
        </p>
    </div>

    <!-- 추가 이미지 -->
    <div style="margin: 40px 0;">
        {"".join([f'<img src="{img}" style="max-width: 100%; margin: 10px 0;" alt="상품 이미지">' for img in candidate.source_images[1:4]])}
    </div>

    <!-- 구매 안내 -->
    <div style="background: #E3F2FD; padding: 20px; margin: 20px 0; border-radius: 8px;">
        <h2 style="color: #1565C0;">📦 배송 안내</h2>
        <ul style="line-height: 2;">
            <li>배송비: 무료</li>
            <li>배송기간: 영업일 기준 2-3일</li>
            <li>교환/반품: 수령 후 7일 이내</li>
        </ul>
    </div>

</div>
"""
        return html.strip()

    def _generate_memo(self, candidate: SourcingCandidate) -> str:
        """관리용 메모 생성

        Args:
            candidate: 소싱 후보

        Returns:
            str: 메모 내용
        """
        lines = [
            f"키워드: {candidate.keyword}",
            f"원가(CNY): ¥{candidate.source_price_cny:.0f}",
            f"원가(KRW): ₩{candidate.estimated_cost_krw:,.0f}",
            f"손익분기: ₩{candidate.breakeven_price:,}",
            f"네이버 최저가: ₩{candidate.naver_min_price:,}",
            f"네이버 평균가: ₩{candidate.naver_avg_price:,}",
            f"경쟁사: {candidate.competitor_count}개",
            f"리스크: {candidate.risk_level.value}",
        ]

        if candidate.risk_reasons:
            lines.append(f"리스크사유: {', '.join(candidate.risk_reasons)}")

        return "\n".join(lines)

    def _adjust_column_widths(self, ws):
        """열 너비 자동 조정"""
        widths = {
            1: 40,   # 상품명
            2: 12,   # 판매가
            3: 12,   # 원가
            4: 10,   # 마진율
            5: 15,   # 카테고리
            6: 10,   # 상품상태
            7: 12,   # 배송비유형
            8: 10,   # 배송비
            9: 10,   # 재고
            10: 50,  # 대표이미지
            11: 50,  # 추가이미지
            12: 100, # 상세페이지
            13: 50,  # 1688 URL
            14: 30,  # 메모
        }

        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    def export_for_review(self, filename: str = None) -> str:
        """대기 중인 상품을 검토용으로 추출 (상세페이지 없이)

        Args:
            filename: 출력 파일명

        Returns:
            str: 생성된 엑셀 파일 경로
        """
        candidates = self.repository.get_candidates_by_status(CandidateStatus.PENDING)

        if not candidates:
            print("[ExcelExporter] 대기 중인 상품이 없습니다.")
            return None

        print(f"[ExcelExporter] {len(candidates)}개 상품 검토용 추출 중...")

        wb = Workbook()
        ws = wb.active
        ws.title = "검토용"

        # 검토용 컬럼 (간략화)
        review_columns = [
            "ID", "상품명", "원가(CNY)", "원가(KRW)", "추천판매가",
            "마진율", "손익분기", "네이버최저가", "경쟁사수", "리스크", "1688 URL"
        ]

        # 헤더
        for col, header in enumerate(review_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 데이터
        for row, c in enumerate(candidates, 2):
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=c.title_kr)
            ws.cell(row=row, column=3, value=c.source_price_cny)
            ws.cell(row=row, column=4, value=int(c.estimated_cost_krw))
            ws.cell(row=row, column=5, value=c.recommended_price)
            ws.cell(row=row, column=6, value=f"{c.estimated_margin_rate:.0%}")
            ws.cell(row=row, column=7, value=c.breakeven_price)
            ws.cell(row=row, column=8, value=c.naver_min_price)
            ws.cell(row=row, column=9, value=c.competitor_count)
            ws.cell(row=row, column=10, value=c.risk_level.value)
            ws.cell(row=row, column=11, value=c.source_url)

        # 파일 저장
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"review_{timestamp}.xlsx"

        filepath = self.output_dir / filename
        wb.save(filepath)

        print(f"[ExcelExporter] 저장 완료: {filepath}")
        return str(filepath)


# CLI 테스트
if __name__ == "__main__":
    exporter = NaverExcelExporter()

    # 승인된 상품 추출
    result = exporter.export_approved()
    if result:
        print(f"엑셀 파일 생성됨: {result}")
    else:
        print("추출할 상품이 없습니다.")
